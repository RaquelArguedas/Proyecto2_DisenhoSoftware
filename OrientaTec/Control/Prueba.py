
"""import pymongo
MONGO_HOST="localhost"
MONGO_PUERTO="27017"
MONGO_TIEMPO_FUERA=1000
MONGO_URI="mongodb://"+MONGO_HOST+":"+MONGO_PUERTO+"/"
MONGO_CLIENTE = None
MONGO_BASEDATOS = None
collecFtProf = None
collecAfiche = None 
collecEvLista = None
collecEvFoto= None

def connectMongoServer():
    try:
        global collecFtProf,collecAfiche,collecEvLista,collecEvFoto,MONGO_CLIENTE, MONGO_BASEDATOS
        MONGO_CLIENTE = pymongo.MongoClient(MONGO_URI,serverSelectionTimeoutMS=MONGO_TIEMPO_FUERA)
        MONGO_BASEDATOS = MONGO_CLIENTE["FotosOrientaTEC"]   
        collecFtProf = MONGO_BASEDATOS["FotosProfesores"]
        collecAfiche = MONGO_BASEDATOS["AficheActividad"]
        collecEvLista = MONGO_BASEDATOS["EvListaAsistencia"]
        collecEvFoto= MONGO_BASEDATOS["EvFotosParticipantes"]
        print("Conexion a Mongo exitosa.")
    except Exception as ex:
        print(ex)

def closeMongoConnection():
    try:
        MONGO_CLIENTE.close()
    except Exception as ex:
        print(ex)
#-----------MÉTODOS PARA BUSCAR FOTOS DE MONGO--------------------------#
#Registrar NUEVOS archivos 
def registrarFotoProfesor(idProfe,nuevaFoto):
    try:
        connectMongoServer()
        #revisar que no exista el str(registro) 
        cantstr(Registro)s = collecFtProf.count_documents({'idProfe':idProfe})
        if(cantstr(Registro)s > 0):
            print("El str(registro) ya existe, NO SE PUEDE actualizar.")
        else:
            collecFtProf.insert_one({'idProfe':idProfe, 'foto': nuevaFoto})
            print("Se ha insertado la foto exitosamente.")
        closeMongoConnection()
    except Exception as ex:
        print(ex)

def registrarFotoAfiche(idActividad,nuevaFoto):
    try:
        connectMongoServer()
        collecAfiche.insert_one({'idActividad':idActividad, 'foto': nuevaFoto})
        print("Foto registrada con exito.")
        closeMongoConnection()
    except Exception as ex:
        print(ex)

def registrarFotoEvLista(idEvidencia,nuevaFoto):
    try:
        connectMongoServer()
        collecEvLista.insert_one({'idEvidencia':idEvidencia, 'foto': nuevaFoto})
        print("Foto registrada con exito.")
        closeMongoConnection()
    except Exception as ex:
        print(ex)

def registrarFotoEv(idEvidencia,nuevaFoto):
    try:
        connectMongoServer()
        collecEvFoto.insert_one({'idEvidencia':idEvidencia, 'foto': nuevaFoto})
        print("Foto registrada con exito.")
        closeMongoConnection()
    except Exception as ex:
        print(ex)
#-------------------------------REGISTRAR-------------------------------

#-------------------------------SET-------------------------------
def setFotoProfesor(idBuscado, nuevaFoto):
    try:
        connectMongoServer()
        #Revisar que exista el str(registro) 
        cantstr(Registro)s = collecFtProf.count_documents({'idProfe':idBuscado})
        if(cantstr(Registro)s > 0):
            collecFtProf.update_one({'idProfe':idBuscado},{'$set':{'foto':nuevaFoto}})
            print("Se ha actualizado la foto exitosamente. ")
        else:
            print("El str(registro) que intenta actualizar NO existe.")
        closeMongoConnection()
    except Exception as ex:
        print(ex)
"""
"""
    El profesor es el único que necesita actualizar str(registro)s porque mantiene UNA única foto, 
    pueden existir n fotos de una evidencia, n fotos de un afiche para una actividad. 
"""
"""
#-------------------------------SET-------------------------------

#-------------------------------GETTERS-------------------------------
def getFotoProfesor(idBuscado):
    try:
        connectMongoServer()
        #revisar que el profesor exista
        cantstr(Registro)s = collecFtProf.count_documents({'idProfe':idBuscado})
        if cantstr(Registro)s > 0 :
            for documento in collecFtProf.find({"idProfe": idBuscado},{ "idProfe": 0, "_id":0}):
                return documento["foto"]
            closeMongoConnection()
        else:
            print("El str(registro) que intenta obtener NO existe.")

    except Exception as ex:
        print(ex)

def getFotoAfiche(idBuscado):
    try:
        connectMongoServer()
        #revisar que el str(registro) exista
        cantstr(Registro)s = collecAfiche.count_documents({'idActividad':idBuscado})
        if cantstr(Registro)s > 0 :
            for documento in collecAfiche.find({"idActividad": idBuscado},{ "idActividad": 0, "_id":0}):
                return documento["foto"]
            closeMongoConnection()
        else:
            print("La actividad que busca NO existe.")
    except Exception as ex:
        print(ex)

def getFotoEvLista(idBuscado):
    try:
        connectMongoServer()
        #revisar que el str(registro) exista
        cantstr(Registro)s = collecEvLista.count_documents({'idActividad':idBuscado})
        if cantstr(Registro)s > 0 :
            for documento in collecEvLista.find({"idEvidencia": idBuscado},{ "idEvidencia": 0, "_id":0}):
                return documento["foto"]
            closeMongoConnection()
        else:
            print("La evidencia que busca NO existe.")
    except Exception as ex:
        print(ex) 

def getFotoEv(idBuscado):
    try:
        connectMongoServer()
        cantstr(Registro)s = collecEvFoto.count_documents({'idActividad':idBuscado})
        if cantstr(Registro)s > 0 :
            for documento in collecEvFoto.find({"idEvidencia": idBuscado},{ "idEvidencia": 0, "_id":0}):
                return documento["foto"]
            closeMongoConnection()
        else:
            print("La evidencia que busca NO existe.")
    except Exception as ex:
        print(ex)
#-------------------------------GETTERS-------------------------------

#setFotoProfesor(1, "Puppy Mini")
#setFotoProfesor(3, "Puppy Mini")
#print(getFotoProfesor(3))
"""
#PRUEBAS EXCEL 

from openpyxl import Workbook
from openpyxl import load_workbook
estudiantes = [['222','Juan','Perez','Brenes','134','jjj@hot',1],['222','Jose','Perez','Brenes','134','jjj@hot',2]]

#Sede 1 SJ, 2 CA, 3 SC, 4 AL, 5 LI
#Si la sede tiene un valor distinto se genera un excel de todos estudiantes
'''params
     @sede: Numero de sede de la cual quiere el excel'''
def generarExcel(sede):
    #crear un nuevo excel 
    wb = Workbook() # se crea el nuevo xlsx
    ws1 = wb['Sheet']  #primer hoja del excel
    #Caso: Excel TODAS sedes
    if(sede != 1 and sede != 2 and sede != 3 and sede != 4 and sede != 5):
        #Necesito hacer 5 hojas en el excel
        ws1.title = 'SJ' #Por defecto la primer hoja que se creacion el libro se llama asi
        ws2 = wb.create_sheet('CA')
        ws3 = wb.create_sheet('SC')
        ws4 = wb.create_sheet('AL')
        ws5 = wb.create_sheet('LI')
        #Headers de los archivos 
        wb['SJ']['A1'] = 'Carne'
        wb['CA']['A1'] = 'Carne'
        wb['SC']['A1'] = 'Carne'
        wb['AL']['A1'] = 'Carne'
        wb['LI']['A1'] = 'Carne'

        wb['SJ']['B1'] = 'Nombre'
        wb['CA']['B1'] = 'Nombre'
        wb['SC']['B1'] = 'Nombre'
        wb['AL']['B1'] = 'Nombre'
        wb['LI']['B1'] = 'Nombre'

        wb['SJ']['C1'] = 'Apellido1'
        wb['CA']['C1'] = 'Apellido1'
        wb['SC']['C1'] = 'Apellido1'
        wb['AL']['C1'] = 'Apellido1'
        wb['LI']['C1'] = 'Apellido1'

        wb['SJ']['D1'] = 'Apellido2'
        wb['CA']['D1'] = 'Apellido2'
        wb['SC']['D1'] = 'Apellido2'
        wb['AL']['D1'] = 'Apellido2'
        wb['LI']['D1'] = 'Apellido2'    

        wb['SJ']['E1'] = 'NumeroCelular'
        wb['CA']['E1'] = 'NumeroCelular'
        wb['SC']['E1'] = 'NumeroCelular'
        wb['AL']['E1'] = 'NumeroCelular'
        wb['LI']['E1'] = 'NumeroCelular'

        wb['SJ']['F1'] = 'CorreoElectronico'
        wb['CA']['F1'] = 'CorreoElectronico'
        wb['SC']['F1'] = 'CorreoElectronico'
        wb['AL']['F1'] = 'CorreoElectronico'
        wb['LI']['F1'] = 'CorreoElectronico' 

        wb['SJ']['G1'] = 'Sede'
        wb['CA']['G1'] = 'Sede'
        wb['SC']['G1'] = 'Sede'
        wb['AL']['G1'] = 'Sede'
        wb['LI']['G1'] = 'Sede'
        #Headers de los archivos 
         
         #Se recorren los estudiantes y se van guardando 
        '''
        for i in range(len(estudiantes)):
            registro = i+2
            sede = None
            if (estudiantes[i][6] == 1):
                sede = 'SJ'
            else:
                sede = 'CA'
            
            wb[sede]['A'+ str(registro)] = estudiantes[i][0]  
            wb[sede]['B'+ str(registro)] = estudiantes[i][1] 
            wb[sede]['C'+ str(registro)] = estudiantes[i][2]
            wb[sede]['D'+ str(registro)] = estudiantes[i][3]
            wb[sede]['E'+ str(registro)] = estudiantes[i][4]
            wb[sede]['F'+ str(registro)] = estudiantes[i][5]
            wb[sede]['G'+ str(registro)] = estudiantes[i][6]
            '''
        for est in estudiantes:
            ws1.append[est]    
    else:
        pass
    wb.save('listaEstudiantes.xlsx')

#recibir un archivo
def cargarExcel(nombArchivo):
    wb = load_workbook(nombArchivo) 
    sheet = wb.active
    i = 2
    listEst = [] #Lista que guarda todos losregistros
    #Recorre cada fila del excel
    for row in sheet.iter_rows(min_row = 2,min_col=1):
        estudiante = [] #Lista que va a guardar los valores del estudiante
        for cell in row:
            if cell.value == None:
                estudiante = [] #se vuelve a poner la lista en vacío
                break #Ese estudiante NO se agrega 
            elif cell.value != None:
                #Se agrega al objeto estudiante
                estudiante.append(cell.value)
        if estudiante != []:
            listEst.append(estudiante) 
        #Se terminó de recorrer el registro(fila)
    print(listEst)


generarExcel(400)
#cargarExcel('C:/Users/Harrick Mc Lean M/OneDrive/Confirmación/Junta Higuito/ignorar.xlsx')
#cargarExcel('ignorar.xlsx')