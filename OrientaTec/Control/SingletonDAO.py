# basado en codigo de Refactoring.Guru, adjuntamos el enlace a continuacion
# https://refactoring.guru/design-patterns/singleton/python/example#:~:text=Singleton%20is%20a%20creational%20design,the%20modularity%20of%20your%20code.

import mysql.connector
import pymongo 
from operator import attrgetter
import random

from datetime import datetime, date, timedelta
from openpyxl import Workbook #Para manejo de excel 
from openpyxl import load_workbook

import os
from pathlib import Path
from operator import attrgetter

#from google.oauth2 import service_account
#from googleapiclient.discovery import build
#from googleapiclient.http import MediaFileUpload

import sys
#Anexo el Directorio en donde se encuentra la clase a llamar
sys.path.append('./Modelo')

#sys.path.append('./Modelo')
#Importo la Clase
from Usuario import *
from Estudiante import * 
from EquipoGuia import *
from Actividad import *
from Profesor import *
from AsistenteAdministrativo import *
from Bitacora import *
from PlanTrabajo import *
from Recordatorio import *
from Evidencia import *
from Observacion import *
from Comentario import *
from Sede import *
from Ordenamiento import *
from Notificacion import *


class SingletonMeta(type):

    _instances = {}

    def __call__(cls, *args, **kwargs):

        if cls not in cls._instances:  #metodo obligatorio getInstance
            instance = super().__call__(*args, **kwargs)
            cls._instances[cls] = instance
        return cls._instances[cls]


class SingletonDAO(metaclass=SingletonMeta):
    
    #Atributos
    #Atributos de conexión
    connection = None
    cursor = None
    #Atributos para conetarse a GOOGLE DRIVE
    # Ruta al archivo JSON de las credenciales de Google Drive
    CREDENCIALES_JSON = 'C:/Users/Harrick Mc Lean M/Downloads/disehnosoftware-50f7c4a809ca.json'

    # Crea una instancia del cliente de Google Drive
    #credenciales = service_account.Credentials.from_service_account_file(CREDENCIALES_JSON)
    #servicio_drive = build('drive', 'v3', credentials=credenciales)


    #Atributos para conetarse a MONGO
    MONGO_HOST="localhost"
    MONGO_PUERTO="27017"
    MONGO_TIEMPO_FUERA=1000
    MONGO_URI="mongodb://"+MONGO_HOST+":"+MONGO_PUERTO+"/"
    MONGO_CLIENTE = None
    MONGO_BASEDATOS = None
    collecFtProf = None
    collecFtAs = None
    collecAfiche = None 
    collecEvLista = None
    collecEvFoto= None
    #Atributos del modelo
    usuarios = []
    estudiantes = []
    equiposGuia = []
    actividades = []
    profesores = []
    asistentes = []
    bitacoras = []
    planesTrabajo = []
    recordatorios = []
    evidencias = []
    observaciones = []
    comentarios = []
    notificaciones = []

    #Constructor que instancia los objetos necesarios del modelo
    def __init__(self):
        self.notificaciones = self.setFromBD("SELECT * from ", "Notificacion")
        self.usuarios = self.setFromBD("SELECT * from ", "Usuario")
        self.estudiantes = self.setFromBD("SELECT * from ", "Estudiante")
        self.bitacoras = self.setFromBD("SELECT * from ", "Bitacora")
        self.profesores = self.setFromBD("SELECT * from ", "Profesor")
        self.equiposGuia = self.setFromBD("SELECT * from ", "EquipoGuia")
        self.recordatorios = self.setFromBD("SELECT * from ", "Recordatorio")
        self.actividades = self.setFromBD("SELECT * from ", "Actividad")
        self.asistentes = self.setFromBD("SELECT * from ", "AsistenteAdministrativo")
        self.planesTrabajo = self.setFromBD("SELECT * from ", "PlanTrabajo")
        self.evidencias = self.setFromBD("SELECT * from ", "Evidencia")
        self.observaciones = self.setFromBD("SELECT * from ", "Observacion")
        self.comentarios = self.setFromBD("SELECT * from ", "Comentario")
    
    #Auxiliar del constructor 
    def setFromBD(self, command, tablaBD):
        self.connectServer()

        try:
            self.cursor.execute(command + tablaBD)

            salida = self.cursor.fetchall()
            
            lista = []
            objeto = []
            for row in salida:
                for item in row:
                    objeto += [item]
                lista += [self.generarObjeto(tablaBD, objeto)]
                objeto = []
            return lista
            
            
        except Exception as ex:
            print(ex)

        self.closeConnection()

    #Auxiliar del constructor
    def generarObjeto(self, tablaBD, lista):
        objeto = None
        
        if (tablaBD == "Usuario"):
            objeto = Usuario(lista[0], lista[1], lista[2], lista[3], lista[4], lista[5], lista[6], self.getNotificaciones(lista[0]))
        elif (tablaBD == "Estudiante"):
            objeto = Estudiante(lista[0], lista[1], lista[2], lista[3], lista[4], lista[5], lista[6], lista[7])
        elif (tablaBD == "EquipoGuia"):
            objeto = EquipoGuia(lista[1], self.generarBitacorasEquipoGuia(lista[0]), self.generarProfesores(str(lista[0])), lista[2])
        elif (tablaBD == "Actividad"):
            objeto = Actividad.Actividad(lista[0], lista[1],lista[2], lista[3], lista[4],lista[5], self.generarResposables(lista[0]), self.generarRecordatorios(lista[0]), lista[6], lista[7], lista[8],lista[9], self.generarBitacorasActividades(lista[0]))
        elif (tablaBD == "Profesor"):
            objeto = Profesor(self.generarCodigoProfesor(lista[5],lista[0]), lista[0],lista[1],lista[2],lista[3], lista[4], lista[5], lista[6],lista[7], lista[8], lista[9], lista[10])
        elif (tablaBD == "AsistenteAdministrativo"):
            objeto = AsistenteAdministrativo(lista[0], lista[1], lista[2], lista[3], lista[4], lista[5], lista[6], lista[7], lista[8])
        elif (tablaBD == "Bitacora"):
            objeto = Bitacora(lista[0], lista[1], lista[2], lista[3], lista[4])
        elif (tablaBD == "PlanTrabajo"):
            objeto = PlanTrabajo(lista[0], lista[1], self.obtenerActividadesPlan(lista[0]))
        elif (tablaBD == "Recordatorio"):
            objeto = Recordatorio(lista[1], lista[2])
        elif (tablaBD == "Evidencia"):
            objeto = Evidencia(lista[1], lista[2])
        elif (tablaBD == "Observacion"):
            objeto = Observacion(lista[0], lista[1], lista[2])
        elif (tablaBD == "Comentario"):
            objeto = Comentario(lista[1], lista[2], lista[3], lista[4], lista[5], lista[0])
        elif (tablaBD == "Notificacion"):
            objeto = Notificacion(lista[0], lista[1], lista[2], lista[3], None)

        return objeto

    def getNotificaciones(self, idUsuario):
        self.connectServer()

        try:
            self.cursor.execute("select * from NotificacionXUsuario where idUsuario = " + str(idUsuario))

            salida = self.cursor.fetchall()

            lista = []
            for row in salida:
                lista += [self.getNotificacion(row[1], idUsuario)]
            return lista
            
            
        except Exception as ex:
            print(ex)

        self.closeConnection()
        
    def getNotificacion(self, idNotificacion, idUsuario):
        for noti in self.notificaciones:
            if(int(noti.idNotificacion) == int(idNotificacion)):
                #devuelve una nueva instancia para cada notificacion del usuario
                return Notificacion(noti.idNotificacion, noti.emisor, noti.fechaHora, noti.contenido, self.getLeida(idNotificacion, idUsuario))

    def getLeida(self, idNotificacion, idUsuario):
        res = self.executeStoredProcedure('getLeida', [idNotificacion, idUsuario])
        return bool(res[0])

    def generarProfesores(self, idEquipoGuia):
        self.connectServer()

        try:
            self.cursor.execute("select profesor.idProfesor from profesoresxequipoguia  inner join profesor on profesor.idProfesor = profesoresxequipoguia.idProfesor where idEquipoGuia = " + str(idEquipoGuia))

            salida = self.cursor.fetchall()

            lista = []
            for row in salida:
                lista += [self.getProfesor(row[0])]
            return lista
            
            
        except Exception as ex:
            print(ex)

        self.closeConnection()

    def generarCodigoProfesor(self, idSede, id):
        return Sede(idSede).name +"-"+str(id)
    
    def getProfesorCedula(self, cedula):
        for prof in self.profesores:
            if(int(prof.cedula) == int(cedula)):
                return prof

    def getAllProfesores(self):
        return self.profesores

    def generarResposables(self, idActividad):
        self.connectServer()

        try:
            self.cursor.execute("select profesor.idProfesor from responsablexactividad  inner join profesor on profesor.idProfesor = responsablexactividad.idResponsable where idActividad = " + str(idActividad))

            salida = self.cursor.fetchall()

            lista = []
            for row in salida:
                lista += [self.getProfesor(row[0])]

            return lista
            
            
        except Exception as ex:
            print(ex)

        self.closeConnection()

    def generarRecordatorios(self, idActividad):
        lista = []
        for recordatorio in self.recordatorios:
            if (recordatorio.idActividad == idActividad):
                lista += [recordatorio]
        return lista

    def obtenerActividadesPlan(self, idPlan):
        self.connectServer()

        try:
            self.cursor.execute("select idActividad from actividadesxplan where idPlan = " + str(idPlan))

            salida = self.cursor.fetchall()

            lista = []
            for row in salida:
                lista += [self.verActividad(row[0])]

            return lista
            
            
        except Exception as ex:
            print(ex)

        self.closeConnection()

    def generarBitacorasEquipoGuia(self, idEquipoGuia):
        self.connectServer()

        try:
            self.cursor.execute("select idBitacora from bitacoraxequipoguia where idEquipoGuia = " + str(idEquipoGuia))

            salida = self.cursor.fetchall()

            lista = []
            for row in salida:
                lista += [self.getBitacora(row[0])]

            return lista
            
            
        except Exception as ex:
            print(ex)

        self.closeConnection()

    def generarBitacorasActividades(self, idActividad):
        self.connectServer()

        try:
            self.cursor.execute("select idBitacora from bitacoraxactividades where idActividad= " + str(idActividad))

            salida = self.cursor.fetchall()

            lista = []
            for row in salida:
                lista += [self.getBitacora(row[0])]
            return lista
            
            
        except Exception as ex:
            print(ex)

        self.closeConnection()

    def getBitacora(self, idBitacora):
        for i in range (len(self.bitacoras)):
            if (self.bitacoras[i].idBitacora == idBitacora):
                return self.bitacoras[i]
        return None 

    #Métodos
    #función que realiza la conexión a MySQL
    def connectServer(self):
        try:
            self.connection = mysql.connector.connect(
                #host = 'localhost',
                host = '127.0.0.1',
                port = 3306,
                user = 'root',
                password = 'abd00123',
                db = 'orientatec'
            )
            if self.connection.is_connected():
                self.cursor = self.connection.cursor()
                #print("Connection succesful")
        except Exception as ex:
            print(ex)


    #función que cierra la conexión a MySQL
    def closeConnection(self):
        if self.connection!=None:
            self.connection.close()
            #print("Connection closed")

    
    #+getActividades():Collection<Actividad>
    def getActividades(self):
        print(self.planesTrabajo[-1].actividades)
        print(self.planesTrabajo[-1].anno)
        
        #listaSalida = sorted(self.planesTrabajo[-1].actividades, key=attrgetter('fechaActividad'))
        #return listaSalida
        return self.planesTrabajo[-1].actividades
    
    #+getTodasActividades():Collection<Actividad>
    def getTodasActividades(self):
        listaSalida = []
        for ac in self.actividades:
            if (self.actividadPertenecePlan(ac) == False):
                listaSalida += [ac]
        print(listaSalida)
        listaSalida = sorted(listaSalida, key=attrgetter('fechaActividad'))
        return listaSalida
    
    def actividadPertenecePlan(self, actividad):
        for plan in self.planesTrabajo:
            for ac in plan.actividades:
                if (ac == actividad):
                    return True
        return False
    
    # +getPlanTrabajo():planTrabajo: PlanTrabajo  
    def getPlanTrabajo(self):
        return self.planesTrabajo[-1]

    # +getConformacionEquipoGuía():Collection<Profesores>
    def getConformacionEquipoGuia(self):
        return self.equiposGuia[-1].listaProfesores 
    
    def getAllProfesores(self):
        return self.profesores
    
    def getAllAsistentes(self):
        return self.asistentes

    # +agregarProfesor(profesor: Profesor): boolean
    def agregarProfesor(self, profesor):
        idEquipoGuia = len(self.equiposGuia)
        
        args = [profesor.id, idEquipoGuia]

        #se agrega a la bd
        id = self.executeStoredProcedure('createprofesoresxequipoguia', args)
        if(len(id)==1):
            #se agrega al equipo guia del año actual el profesor
            self.equiposGuia[idEquipoGuia-1].agregarProfesor(profesor)
        
        return id
    
    # +verActividad(id): Actividad
    def verActividad(self, id):
        for i in range (len(self.actividades)):
            if (int(self.actividades[i].idActividad) == int(id)):
                return self.actividades[i]
        return None 
    
    #devuelve los comentario de una actividad
    def verComentariosActividad(self, idActividad):
        lista = []

        for i in range (len(self.comentarios)):
            if (self.comentarios[i].idActividad == idActividad):
                lista += [self.comentarios[i]]
        return lista 

    #devuelve la evidencia de una actividad
    def verEvidenciasActividad(self, idActividad):
        lista = []
        for i in range (len(self.evidencias)):
            if (self.evidencias[i].idActividad == idActividad):
                lista += [self.evidencias[i]]
        return lista

    # +agregarActividad(actividad: Actividad): boolean
    def agregarActividad(self, actividad, idPlan):
        
        args = [actividad.idActividad, idPlan]

        #se agrega a la bd
        id = self.executeStoredProcedure('createActividadesxPlan', args)
        
        if(len(id)==1):
            #se agrega al plan la actividad
            for plan in self.planesTrabajo:
                if (plan.idPlan == idPlan):
                    plan.agregarActividad(actividad)
        
        return id
    
    def crearPlanTrabajo(self, anno):
        
        args = [anno]

        #se agrega a la bd
        id = self.executeStoredProcedure('createPlanTrabajo', args)
        
        if(len(id)==1):
            #se obtiene el id y se le agrega
            salida = PlanTrabajo(id[0], anno, []) 

            #se agrega a la lista de planes
            self.planesTrabajo += [salida]
        
        return id

    # +consultarProximaActividad: actividad: Actividad
    def consultarProximaActividad(self):
        # Creando lista de fechas
        fechas = []
        for actividad in self.actividades:
            h = actividad.horaInicio
            fechas += [datetime.combine(actividad.fechaActividad, (datetime.min + h).time())]
                    
        # busca la fecha mas cercana a la actual
        cloz_dict = {
        d.timestamp() - datetime.now().timestamp() : d
        for d in fechas}

        # encontrando la fecha mas cercana
        min = 9000000000000000
        for key in cloz_dict:
            if (key>=0 and key<min):
                min = key

        #obtiene el indice de la actividad que tiene la fecha mas cercana
        index = list(cloz_dict.keys()).index(min)
        
        return self.actividades[index]

    # +modificarActividad(data): boolean
    def modificarActividad(self, idActividad, nombreActividad,
                            tipoActividad, fechaActividad, horaInicio,
                            horaFin, recordatorio, medio, enlace,
                            estado):
        
        ultimaModificacion = date.today()
        
        args = [idActividad, nombreActividad, tipoActividad, fechaActividad, horaInicio,
                horaFin, medio, enlace, estado, ultimaModificacion]
        print(args)

        #se modifica en la bd
        respuesta = self.executeStoredProcedure('updateActividad', args)
        print(idActividad, type(idActividad))
        #Si no hubo errores dentro de la BD, realiza las modificaciones correspondientes
        if (respuesta == None):
            print("entr a res")
            #se modifica en lista
            for i in range(len(self.actividades)):
                if (self.actividades[i].idActividad == idActividad):
                    if (nombreActividad != None):
                        self.actividades[i].nombreActividad = nombreActividad
                    if (tipoActividad != None):
                        self.actividades[i].tipoActividad = tipoActividad
                    if (fechaActividad != None):
                        self.actividades[i].fechaActividad = fechaActividad
                    if (horaInicio != None):
                        self.actividades[i].horaInicio = horaInicio
                    if (horaFin != None):
                        self.actividades[i].horaFin = horaFin
                    if (recordatorio != None):
                        listaRecordatorios = []
                        for fecha in recordatorio:
                            listaRecordatorios.append(Recordatorio(idActividad, fecha))
                        self.updateRecordatorios(idActividad, listaRecordatorios) #actualiza la lista de recordatorios y la bd
                        self.actividades[i].recordatorios = listaRecordatorios #actualiza el objeto
                    if (medio != None):
                        self.actividades[i].medio = medio
                    if (enlace != None):
                        self.actividades[i].enlace = enlace
                    if (estado != None):
                        self.actividades[i].estado = estado
                        print("estado modificado", estado)
                    if (ultimaModificacion != None):
                        self.actividades[i].ultimaModificacion = ultimaModificacion
                
        return respuesta
    
    def updateRecordatorios(self, idActividad, recordatorios):
        #eliminar todos los recordatorios con ese idActividad
        self.executeStoredProcedure('deleteRecordatorioActividad', [int(idActividad)])
        #Crear los recordatorios
        for rec in recordatorios:
            print("cree")
            self.executeStoredProcedure('createRecordatorio', [int(idActividad), rec.fecha])
        #actualizar la lista
        for rec in self.recordatorios:
            if rec.idActividad == idActividad:
                print("quite")
                self.recordatorios.remove(rec)
        self.recordatorios += recordatorios

        

    def agregarResponsablesActividad(self, idActividad, responsablesNuevos):
        actividad = None
        #se obtiene la actividad
        for ac in self.actividades:
            if (ac.idActividad == idActividad):
                actividad = ac

        for responsable in responsablesNuevos:
            #agregar en tabla responsablexactividad
            self.executeStoredProcedure("createResponsableXActividad", [int(responsable['id']), int(idActividad)])

            #agregar a la actividad el responsable
            print("DESDE AGREGAR RESPONSABLE", actividad.nombreActividad, actividad.responsables, type(actividad.responsables))
            actividad.agregarResponsable(self.getProfesor(int(responsable['id'])))

    def quitarResponsablesActividad(self, idActividad, responsablesEliminados):
        actividad = None
        #se obtiene la actividad
        for ac in self.actividades:
            if (ac.idActividad == idActividad):
                actividad = ac

        print('Resp. elim: ', responsablesEliminados)
        for responsable in responsablesEliminados:
            #quitar en tabla responsablexactividad
            self.executeStoredProcedure("quitarResponsablesActividad", [responsable, idActividad])

            #quita a la actividad el responsable
            actividad.quitarResponsable(self.getProfesor(responsable))
            
    def bitacoraEquipoGuia(self, fecha, hora, idAutor, descripcion):
        #crear bitacora
        idBitacora = self.executeStoredProcedure("createBitacora", [fecha, hora, idAutor, descripcion])

        if (len(idBitacora) == 1):
            #insertar en la lista del singleton
            bitacora = Bitacora(idBitacora[0], fecha, hora, idAutor, descripcion)

            self.bitacoras += [bitacora]

            #crear bitacoraxequipoguia
            self.executeStoredProcedure("modificarEquipoGuia", [date.today().year, idBitacora[0]])

            #insertar en modificaciones del equipo guia
            self.equiposGuia[-1].agregarModificacion(bitacora)

    def bitacoraActividad(self, idActividad, fecha, hora, idAutor, descripcion):
        for ac in self.actividades:
            print(ac.idActividad, ac.nombreActividad)

        print(idActividad, type(idActividad))
        print("Tipo de la actividad:" , type(self.verActividad(int(idActividad))))
        #crear bitacora
        idBitacora = self.executeStoredProcedure("createBitacora", [fecha, hora, idAutor, descripcion])

        if (len(idBitacora) == 1):
            #insertar en la lista del singleton
            bitacora = Bitacora(idBitacora[0], fecha, hora, idAutor, descripcion)

            print(type(bitacora), bitacora)

            self.bitacoras += [bitacora]

            #crear bitacoraxactividad
            self.executeStoredProcedure("createbitacoraxactividades", [idActividad, idBitacora[0]])

            #insertar en modificaciones de la actividad
            self.verActividad(idActividad).agregarModificacion(bitacora)



    # +cancelarActividad(): boolean, se hizo crear observacion, ya que para modificar el estado de una actividad se hace con modificarActividad()
    def crearObservacion(self, idActividad, fechaCancelacion, detalle):
        
        args = [idActividad, fechaCancelacion, detalle]

        #se agrega a la bd
        id = self.executeStoredProcedure('createObservacion', args)
        
        #si no hubo errores
        if (len(id) == 1):
            #creamos el objeto
            objeto = Observacion(idActividad, fechaCancelacion, detalle) 

            #se agrega a la lista de Observaciones
            self.observaciones += [objeto]
            print(objeto, ", len: ",len(self.observaciones))
        else:
            return id

    # +publicarActividad(data): boolean
    def publicarActividad(self, idActividad):
        respuesta = self.modificarActividad(idActividad, None, None, None, None, None, None,None, None, None, 2, None)

        if (respuesta == None):
            for actividad in self.actividades:
                if (actividad.idActividad == idActividad):
                    actividad.estado = 2

        return respuesta

    # +crearActividad(data): void
    def crearActividad(self, nombreActividad, tipoActividad, fechaActividad,
                horaInicio, horaFin, recordatorio, responsables, medio,  
                enlace, estado):
        
        ultimaModificacion = date.today()

        args = [nombreActividad, tipoActividad, fechaActividad,
                horaInicio, horaFin, medio,  
                enlace, estado, ultimaModificacion]

        #se agrega a la bd
        id = self.executeStoredProcedure('createActividad', args)
        print("id" , id)
        if(len(id)==1):
            print("entro a crearla")
            #se obtiene el id y se le agrega

            listaRecordatorios = []
            for fecha in recordatorio:
                listaRecordatorios.append(Recordatorio(id[0], fecha))

            print(listaRecordatorios)

            salida = Actividad.Actividad(id[0], nombreActividad, tipoActividad, fechaActividad.date(),
                    horaInicio.time(), horaFin.time(),  [], listaRecordatorios, medio,  
                    enlace, estado, ultimaModificacion, [])

            #se agrega a la lista de Actividades
            print(len(self.actividades))
            self.actividades += [salida]
            print(salida.nombreActividad)
            print(len(self.actividades))

            print("responsables: ", responsables)
            self.agregarResponsablesActividad(id[0], responsables)
            print("recordatorios: ", recordatorio)
            self.updateRecordatorios(id[0], listaRecordatorios)
        
        return id


    def crearProfesor(self,cedula,nombre,apellido1, apellido2, sede, numeroCelular,
                        correoElectronico, numeroOficina,autoridad, estado):

        args = [cedula, nombre, apellido1, 
                apellido2, sede, numeroCelular, correoElectronico, 
                numeroOficina, autoridad, estado]

        #se agrega a la bd
        id = self.executeStoredProcedure('createProfesor', args)
        print("id: ", id)
        if(len(id)==1):
            #se obtiene el id y se le agrega
            prof = Profesor(self.generarCodigoProfesor(sede,id[0]), id[0], cedula, nombre, apellido1, 
                apellido2, sede, numeroCelular, correoElectronico, 
                numeroOficina, autoridad, estado) 
            #se agrega a la lista de Actividades
            self.profesores += [prof]

        con = random.randrange(10**7, 10**8)
        rol = 1
        if (autoridad == 1) :
            rol = 2
        self.crearUsuario(correoElectronico, str(con), rol, sede, True, True)
        
        return id

    # +modificarProfesor(idProfesor, ): boolean
    def modificarProfesor(self, id, cedula,nombre,apellido1, apellido2, sede, numeroCelular,
                        correoElectronico, numeroOficina,autoridad, estado):
        
        args = [id,cedula,nombre,apellido1, apellido2, sede, numeroCelular,
                correoElectronico, numeroOficina,autoridad, estado]

        #se modifica en la bd
        respuesta = self.executeStoredProcedure('updateProfesor', args)

        #Si no hubo errores dentro de la BD, realiza las modificaciones correspondientes
        if (respuesta == None):
            #se modifica en lista
            for i in range(len(self.profesores)):
                if (self.profesores[i].id == id):
                    if (cedula != None):
                        self.profesores[i].cedula = cedula
                    if (nombre != None):
                        self.profesores[i].nombre = nombre
                    if (apellido1 != None):
                        self.profesores[i].apellido1 = apellido1
                    if (apellido2 != None):
                        self.profesores[i].apellido2 = apellido2
                    if (sede != None):
                        self.profesores[i].sede = sede
                        self.profesores[i].codigo = self.generarCodigoProfesor(sede,self.profesores[i].id)
                    if (numeroCelular != None):
                        self.profesores[i].numeroCelular = numeroCelular
                    if (correoElectronico != None):
                        self.modificarUsuarioCorreo(self.profesores[i].correoElectronico, correoElectronico)
                        self.profesores[i].correoElectronico = correoElectronico
                    if (numeroOficina != None):
                        self.profesores[i].numeroOficina = numeroOficina
                    if (autoridad != None):
                        self.profesores[i].autoridad = autoridad
                    if (estado != None):
                        self.profesores[i].estado = estado
        return respuesta
    
    # +modificarAsistente(idProfesor, ): boolean
    def modificarAsistente(self, id, cedula,nombre,apellido1, apellido2, sede, numeroCelular,
                        correoElectronico, numeroOficina):
        
        args = [id,cedula,nombre,apellido1, apellido2, sede, numeroCelular,
                correoElectronico, numeroOficina]

        #se modifica en la bd
        respuesta = self.executeStoredProcedure('updateAsistenteAdministrativo', args)

        #Si no hubo errores dentro de la BD, realiza las modificaciones correspondientes
        if (respuesta == None):
            #se modifica en lista
            for i in range(len(self.asistentes)):
                if (self.asistentes[i].id == int(id)):
                    if (cedula != None):
                        self.asistentes[i].cedula = int(cedula)
                    if (nombre != None):
                        self.asistentes[i].nombre = nombre
                    if (apellido1 != None):
                        self.asistentes[i].apellido1 = apellido1
                    if (apellido2 != None):
                        self.asistentes[i].apellido2 = apellido2
                    if (sede != None):
                        self.asistentes[i].sede = int(sede)
                    if (numeroCelular != None):
                        self.asistentes[i].numeroCelular = int(numeroCelular)
                    if (correoElectronico != None):
                        self.modificarUsuarioCorreo(self.asistentes[i].correoElectronico, correoElectronico)
                        self.asistentes[i].correoElectronico = correoElectronico
                    if (numeroOficina != None):
                        self.asistentes[i].numeroOficina = int(numeroOficina)
        return respuesta

    # +darBajaProfesor(idProfesor: int): boolean
    def darBajaProfesor(self, idProfesor):
        
        #borra en la BD que un profesor corresponda al equipo guia actual
        respuesta = self.executeStoredProcedure("darDeBaja", [idProfesor])

        #borar al profesor de la lista del equipo guia actual
        for profesor in self.equiposGuia[-1].listaProfesores:
            if (profesor.id == idProfesor):
                self.equiposGuia[-1].listaProfesores.remove(profesor)

        #cambia el estado del profesor
        self.modificarProfesor(idProfesor, None,None,None, None, None, None, None, None,None, 2)

        return respuesta
    
    # +crearComentario(comentario):void
    def crearComentario(self, idActividad,autor,fechaHora, contenido, idComentarioPadre):
        
        args = [idActividad,autor,fechaHora,contenido, idComentarioPadre]

        #se agrega a la bd
        id = self.executeStoredProcedure('createComentario', args)
        if(len(id)==1):
            #se obtiene el id y se le agrega
            salida = Comentario(idActividad,autor,fechaHora,contenido, idComentarioPadre, id)

            #se agrega a la lista de Actividades
            self.comentarios += [salida]
        
        return id

    # +crearEvidencia(): boolean
    def crearEvidencia(self, idActividad,linkGrabacion):
        args = [idActividad,linkGrabacion]

        #se agrega a la bd
        id = self.executeStoredProcedure('createEvidencia', args)
        
        if(len(id)==1):
            salida = Evidencia(idActividad,linkGrabacion)

            #se agrega a la lista de Actividades
            self.evidencias += [salida]
        
        return id

    # +getEvidencia
    def getEvidencia(self, idActividad):
        for ev in self.evidencias:
            print(ev.idActividad)
            if(int(ev.idActividad) == int(idActividad)):
                return ev

    # +getProfesor(id:int):profesor:Profesor
    def getProfesor(self, idProfesor):
        for prof in self.profesores:
            if(int(prof.id )== int(idProfesor)):
                return prof
            
    def getProfesorCodigo(self, codigo):
        for prof in self.profesores:
            if(prof.codigo == codigo):
                return prof
            
    # +consultarEstudiantes(ordenamiento: enum): Collection<Estudiante>
    def consultarEstudiantes(self):
        return self.estudiantes

    # +buscarEstudiante(id:int) : Estudiante
    def buscarEstudiante(self, carnet):
        for estudiante in self.estudiantes:
            if (int(carnet) == estudiante.carnet):
                print("llegue")
                return estudiante
        return None

    # +modificarEstudiante(estudiante Estudiante): boolean
    def modificarEstudiante(self, carnet, nombre,apellido1, apellido2, sede, correoElectronico, numeroCelular, estado):
        
        args = [carnet, nombre, apellido1, apellido2, sede,  numeroCelular,
                correoElectronico, estado]
        
        #se modifica en la bd
        respuesta = self.executeStoredProcedure('updateEstudiante', args)

        #Si no hubo errores dentro de la BD, realiza las modificaciones correspondientes
        if (respuesta == None):
            #se modifica en lista

            for i in range(len(self.estudiantes)):
                if (self.estudiantes[i].carnet == int(carnet)):
                    if (carnet != None):
                        self.estudiantes[i].carnet = carnet
                    if (nombre != None):
                        self.estudiantes[i].nombre = nombre
                    if (apellido1 != None):
                        self.estudiantes[i].apellido1 = apellido1
                    if (apellido2 != None):
                        self.estudiantes[i].apellido2 = apellido2
                    if (sede != None):
                        self.estudiantes[i].sede = sede
                    if (correoElectronico != None):
                        self.modificarUsuarioCorreo(self.estudiantes[i].correoElectronico, correoElectronico)
                        self.estudiantes[i].correoElectronico = correoElectronico
                    if (numeroCelular != None):
                        self.estudiantes[i].numeroCelular = numeroCelular
        return respuesta

    #modificarUsuario(data):id
    def modificarUsuario(self, idUsuario, correoElectronico, contrasenha, idRol, idSede, permiteNotis, permiteChats):
        
        args = [idUsuario, correoElectronico, contrasenha, idRol, idSede, permiteNotis, permiteChats]
        
        #se modifica en la bd
        respuesta = self.executeStoredProcedure('updateUsuario', args)

        #Si no hubo errores dentro de la BD, realiza las modificaciones correspondientes
        if (respuesta == None):
            #se modifica en lista

            for i in range(len(self.usuarios)):
                if (self.usuarios[i].idUsuario == int(idUsuario)):
                    if (correoElectronico != None):
                        self.usuarios[i].correo = correoElectronico
                    if (contrasenha != None):
                        self.usuarios[i].contrasenha = contrasenha
                    if (idRol != None):
                        self.usuarios[i].idRol = idRol
                    if (idSede != None):
                        self.usuarios[i].idSede = idSede
                    if (permiteNotis != None):
                        self.usuarios[i].permiteNotis = permiteNotis
                    if (permiteChats != None):
                        self.usuarios[i].permiteChats = permiteChats
        return respuesta

    #+crearUsuario(data):id
    def crearUsuario(self, correoElectronico, contrasenha, idRol, idSede, permiteNotis, permiteChats):
        args = [correoElectronico, contrasenha, idRol, idSede, permiteNotis, permiteChats]

        #se agrega a la bd
        id = self.executeStoredProcedure('createUsuario', args)
        
        if(len(id)==1):
            salida = Usuario(id[0], correoElectronico, contrasenha, idRol, idSede, permiteNotis, permiteChats)

            #se agrega a la lista de Actividades
            self.usuarios += [salida]
        print("Se creo el usuario")
        return id

    #+getUsuario(id):usuario
    def getUsuario(self, idUsuario):
        for user in self.usuarios:
            if(user.idUsuario == idUsuario):
                return user

    #getUsuarios():Collection<Usuario>
    def getUsuarios(self):
        print('DAO, Usuarios: ', self.usuarios)
        return self.usuarios
    
    def modificarUsuarioCorreo(self, correoAnterior, correoNuevo):        
        for user in self.usuarios:
            if (user.correo == correoAnterior):
                self.modificarUsuario(user.idUsuario, correoNuevo, None, None, None, None, None)


    #Notificaciones
    #eliminar notificacion
    def deleteNotificacionUsuario(self, idNotificacion, idUsuario):
        id = self.executeStoredProcedure('deleteNotificacionUsuario', [idNotificacion, idUsuario])        
        return id
    
    #eliminar notificaciones de un usuario
    def deleteNotificacionesUsuario(self, idUsuario):
        id = self.executeStoredProcedure('deleteNotificacionesUsuario', [idUsuario])        
        return id
    
    #marcar como leída o no. Si esta leida la marca como no, y si no esta leida la marca como leida
    #ES PARA UNA NOTIFICAION
    def cambiarLeida(self, idNotificacion, idUsuario):
        #cambia el estado en la BD
        id = self.executeStoredProcedure('toggleLeida', [idNotificacion, idUsuario])   

        if (len(id)==1):
            #cambia el atributo en el objeto de la lista de notificaciones del usuario correspondiente
            for usuario in self.usuarios:
                if (usuario.idUsuario == idUsuario):
                    for noti in usuario.notificaciones:
                        if (noti.idNotificacion == idNotificacion):
                            noti.invertirLeida()

        return id

    def todasLeidas(self, idUsuario, leidas):
        #cambia el estado en la BD
        id = self.executeStoredProcedure('setleidasUsuario', [idUsuario, leidas])   
        print("id: ", id)
        if (len(id)==1):
            #cambia el atributo en el objeto de la lista de notificaciones del usuario correspondiente
            for usuario in self.usuarios:
                if (usuario.idUsuario == idUsuario):
                    for noti in usuario.notificaciones:
                        noti.leida = leidas

        return id
            
    #crear notificacion
    #createNotificacion`(in _emisor int, in _fechaHora datetime, in _contenido varchar(50))
    def createNotificacion(self, idUsuarioEmisor, fechaHora, contenido):
        #se crea en la bd
        id = self.executeStoredProcedure('createNotificacion', [int(idUsuarioEmisor), fechaHora, contenido])   

        #si se consigue crear se agrega a la lista de notificaciones del dao
        if (len(id)==1):
            salida = Notificacion(id[0], int(idUsuarioEmisor), fechaHora, contenido, None)

            self.notificaciones += [salida]

        return id #devuelve el id de la notificacion o el error

    #enviar notificacion a usuariosS
    def notificacionUsuarios(self, idNotificacion, idUsuario):
        #genera el campo idNotificacionXUsuario, siempre tiene como leida false
        id = self.executeStoredProcedure('createNotificacionXUsuario', [int(idNotificacion), int(idUsuario), False])   

        #le ingresa la notificacion al usuario correspondiente
        notificacion = None
        if (len(id) == 1):
            for noti in self.notificaciones:
                if (noti.idNotificacion == int(idNotificacion)):
                    notificacion = noti

            for user in self.usuarios:
                if(user.idUsuario == int(idUsuario)):
                    print("se inserto la noti ", notificacion.idNotificacion)
                    user.notificaciones += [Notificacion(notificacion.idNotificacion, notificacion.emisor, notificacion.fechaHora, notificacion.contenido, False)]

        return id
    
    #ejecuta un procedimiento almacenado y devuelve una lista con resultados
    def executeStoredProcedure(self, storedProcName, args):
        self.connectServer()

        try:
            self.cursor.callproc(storedProcName, args)
            self.connection.commit()
            for result in self.cursor.stored_results():
                out = result.fetchall()
                return out[0]
                
        except Exception as ex:
            print(ex)

        self.closeConnection()

        
    #CONEXION A MONGODB
    def connectMongoServer(self):
        try:
            self.MONGO_CLIENTE = pymongo.MongoClient(self.MONGO_URI,serverSelectionTimeoutMS=self.MONGO_TIEMPO_FUERA)
            self.MONGO_BASEDATOS = self.MONGO_CLIENTE["FotosOrientaTEC"]   
            self.collecFtProf = self.MONGO_BASEDATOS["FotosProfesores"]
            self.collecAfiche = self.MONGO_BASEDATOS["AficheActividad"]
            self.collecEvLista = self.MONGO_BASEDATOS["EvListaAsistencia"]
            self.collecEvFoto= self.MONGO_BASEDATOS["EvFotosParticipantes"]
            self.collecFtAs= self.MONGO_BASEDATOS["FotosAsistentes"]
            print("Conexion a Mongo exitosa.")
        except Exception as ex:
            print(ex)
 
    def closeMongoConnection(self):
        try:
            self.MONGO_CLIENTE.close()
        except Exception as ex:
            print(ex)


    def registrarFotoProfesor(self,idProfe,image):
        try:
            self.connectMongoServer()
            #revisar que no exista el registro 
            cantRegistros = self.collecFtProf.count_documents({'idProfe':idProfe})
            print(cantRegistros)
            if(cantRegistros > 0): #si ya existe lo actualiza
                self.collecFtProf.update_one({'idProfe':idProfe},{'$set':{'foto':image.read()}})
                print("Se ha actualizado la foto exitosamente. ")
            else:
                self.collecFtProf.insert_one({'idProfe':idProfe, 'foto': image.read()})
                print("Se ha insertado la foto exitosamente.")
            self.closeMongoConnection()
        except Exception as ex:
            print(ex)

    def registrarFotoAsistente(self,idAsistente,image):
        try:
            self.connectMongoServer()            
            cantRegistros = self.collecFtAs.count_documents({'idAsistente':idAsistente})
            print(cantRegistros)
            if(cantRegistros > 0): #si existe lo actualiza, si no lo crea
                self.collecFtAs.update_one({'idAsistente':idAsistente},{'$set':{'foto':image.read()}})
                print("Se ha actualizado la foto exitosamente. ")
            else:
                self.collecFtAs.insert_one({'idAsistente':idAsistente, 'foto': image.read()})
                print("Se ha insertado la foto exitosamente.")
            self.closeMongoConnection()
        except Exception as ex:
            print(ex)

    def registrarFotoAfiche(self,idActividad,image):
        try:
            self.connectMongoServer()
            #revisar que no exista el registro 
            cantRegistros = self.collecAfiche.count_documents({'idActividad':idActividad})
            print(cantRegistros)
            if(cantRegistros > 0): #si existe lo actualiza
                self.collecAfiche.update_one({'idActividad':idActividad},{'$set':{'foto':image.read()}})
                print("Se ha actualizado la foto exitosamente. ")
            else:
                self.collecAfiche.insert_one({'idActividad':idActividad, 'foto': image.read()})
                print("Se ha insertado la foto exitosamente.")
            self.closeMongoConnection()
        except Exception as ex:
            print(ex)

    def registrarFotoEvLista(self,idActividad,image):
        try:
            self.connectMongoServer()
            self.collecEvLista.insert_one({'idActividad':idActividad, 'foto': image.read()})
            print("Foto registrada con exito.")
            self.closeMongoConnection()
        except Exception as ex:
            print(ex)

    def registrarFotoEv(self,idActividad,image):
        try:
            self.connectMongoServer()
            cantRegistros = self.collecEvFoto.count_documents({'idActividad':idActividad})
            print(cantRegistros)
            if(cantRegistros > 0): #si existe lo actualiza
                self.collecEvFoto.update_one({'idActividad':idActividad},{'$set':{'foto':image.read()}})
                print("Se ha actualizado la foto exitosamente. ")
            else:
                self.collecEvFoto.insert_one({'idActividad':idActividad, 'foto': image.read()})
                print("Se ha insertado la foto exitosamente.")
            self.closeMongoConnection()
        except Exception as ex:
            print(ex)

    def getFotoProfesor(self,idBuscado):
        try:
            self.connectMongoServer()
            #revisar que el profesor exista
            cantRegistros = self.collecFtProf.count_documents({'idProfe':idBuscado})
            if cantRegistros > 0 :                
                document = self.collecFtProf.find_one({'idProfe': idBuscado})
                self.closeMongoConnection()
                return document['foto']
            else:
                print("El registro que intenta obtener NO existe.")
                return None
        except Exception as ex:
            print(ex)

    def getFotoAsistente(self,idBuscado):
        try:
            self.connectMongoServer()
            #revisar que el profesor exista
            cantRegistros = self.collecFtAs.count_documents({'idAsistente':idBuscado})
            if cantRegistros > 0 :                
                document = self.collecFtAs.find_one({'idAsistente': idBuscado})
                self.closeMongoConnection()
                return document['foto']
            else:
                print("El registro que intenta obtener NO existe.")
                return None
        except Exception as ex:
            print(ex)
    
    def getFotoAfiche(self,idBuscado):
        try:
            self.connectMongoServer()
            #revisar que el registro exista
            cantRegistros = self.collecAfiche.count_documents({'idActividad':idBuscado})
            if cantRegistros > 0 :
                document = self.collecAfiche.find_one({'idActividad': idBuscado})
                self.closeMongoConnection()
                return document['foto']
            else:
                print("El registro que intenta obtener NO existe.")
                return None
        except Exception as ex:
            print(ex)

    #i. Una colección de imágenes con la lista de asistencia,
    def getEvLista(self,idBuscado):
        try:
            self.connectMongoServer()
            #revisar que el registro exista
            cantRegistros = self.collecEvLista.count_documents({'idActividad':idBuscado})
            if cantRegistros > 0 :
                resultados = self.collecEvLista.find({'idActividad': idBuscado})
                listaSalida = []
                for r in resultados:
                    listaSalida += [r['foto']]
                self.closeMongoConnection()
                return listaSalida
            else:
                print("El registro que intenta obtener NO existe.")
                return None
        except Exception as ex:
            print(ex) 

    # ii. Una imagen de los participantes, expositores y estudiantes o bien,
    # screenshots de la reunión en caso de haber sido remota. 
    def getFotoEv(self, idBuscado):
        try:
            self.connectMongoServer()
            cantRegistros = self.collecEvFoto.count_documents({'idActividad':idBuscado})
            if cantRegistros > 0 :
                documento = self.collecEvFoto.find_one({"idActividad": idBuscado})
                self.closeMongoConnection()
                return documento["foto"]
            else:
                print("El registro que intenta obtener NO existe.")
                return None
        except Exception as ex:
            print(ex)
    #-------------------------------GETTERS-------------------------------
    #--------------------EXCEL----------------------------
    #Jalar datos para UNA SEDE especifica
    #La sede que recibe es un idSede
    def generarExcelSede(self,sede):
        listaEstudiantes=[['Carne','Nombre', 'Apellido1','Apellido2','Celular','Correo','Sede','Estado']] 
        for estudiante in self.estudiantes:
            est = []
            #Sede: 1Sj   
            if (int(estudiante.sede) == int(sede)):
                #La info de ese registro se guarda 
                est.append(estudiante.carnet)  #Carne
                est.append(estudiante.nombre) #Nombre
                est.append(estudiante.apellido1) #App 1
                est.append(estudiante.apellido2) #App 2
                est.append(estudiante.numeroCelular) #celular
                est.append(estudiante.correoElectronico) #correo
                est.append(estudiante.sede) #Sede
                est.append(estudiante.estado) #Estado 
                listaEstudiantes.append(est)
        return listaEstudiantes
    
    #cargarExcel
    #Lee los registros de un excel en la base de datos
     #   params: @nombArchivo: nombreDelExcel o ruta
    def cargarExcel(self,nombArchivo):
        wb = load_workbook(nombArchivo) 
        sheet = wb.active
        i = 2
        #Recorre cada fila del excel
        for row in sheet.iter_rows(min_row = 2,min_col=1):
            estudiante = [] #Lista que va a guardar los valores del estudiante
            for cell in row:
                if cell.value == None:
                    estudiante = [] #se vuelve a poner la lista en vacío
                    break #Ese estudiante NO se agrega 
                elif cell.value != None:
                    #Se agrega al objeto estudiante
                    estudiante.append(cell.value)
            if estudiante != []: #ya recorrió todos los campos de una fila 
                #Validar que la SEDE sea válida
                if (int(estudiante[6]) == 1 or int(estudiante[6]) == 2 or int(estudiante[6]) == 3 
                or int(estudiante[6]) == 4 or int(estudiante[6]) == 5):  
                    #Llamar al método agregarEstudiante de BD
                    args = [estudiante[0],estudiante[1],estudiante[2],estudiante[3],estudiante[6],
                    estudiante[4],estudiante[5],1]
                    #se modifica en la bd
                    respuesta = self.executeStoredProcedure('createEstudiante', args)
                    print('respuesta storeProcedure:')
                    print(respuesta)
                    if(len(respuesta)==1):
                        #se genera el Estudiante y se agrega a la lista de estudiantes
                        est = Estudiante(estudiante[0],estudiante[1],estudiante[2],estudiante[3],estudiante[6],
                        estudiante[4],estudiante[5],1) 
                        #se agrega a la lista de Estudiantes
                        self.estudiantes += [est]
                        
        return True
    
    #Metodo para agregar estudiantes copnectándose a BD por si necesita en el futuro   
    # +agregarEstudiante(estudiante: Estudiante): boolean
    def agregarEstudiante(self, carnet, nombre,apellido1,
        apellido2, sede, numeroCelular,correoElectronico, estado):

            args = [carnet,nombre,apellido1,apellido2,sede,numeroCelular,correoElectronico,estado]

            #se agrega a la bd
            id = self.executeStoredProcedure('createEstudiante', args)
            if(len(id)==1):
                #se genera el Estudiante y se agrega a la lista de estudiantes
                est = Estudiante(carnet,nombre,apellido1,apellido2,sede,numeroCelular,correoElectronico,estado) 
                #se agrega a la lista de Estudiantes
                self.estudiantes += [est]
            return id
    
    #Funciones que no se copiaron en merge
    def getProfesorCodigo(self, codigo):
        for prof in self.profesores:
            if(prof.codigo == codigo):
                return prof
            
    def getProfesorCedula(self, cedula):
        for prof in self.profesores:
            if(prof.cedula == cedula):
                return prof

